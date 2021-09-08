"""Some helpful functions for file importing, data cleaning/validation, etc."""

import os
import re
import logging
import yaml
import xlrd
import openpyxl
from urllib.parse import urlparse, urlunparse
from pathlib import Path
from datetime import datetime

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June',
          'July', 'August', 'September', 'October', 'November', 'December']
MONTHS_DICT = {i+1: x for (i, x) in enumerate(MONTHS)}

# =============================================================================
# File import/export and naming
# =============================================================================
def import_yaml(path):
    """Import a YAML configuration file as dictionary"""
    with open(path, 'r') as stream:
        try:
            d = yaml.safe_load(stream)
            return d
        except yaml.YAMLError as e:
            print(e)
            return None


# =============================================================================
# String cleaning
# =============================================================================
def to_bool(x):
    """Return True if truthy text, False if falsy text, and None otherwise"""
    s = str(x).lower()
    if s in ['yes', 'y', 'true', 't']:
        return True
    elif s in ['no', 'n', 'false', 'f']:
        return False
    elif s in ['', 'nan', 'none']:
        return None
    else:
        logging.warning(f"Could not convert {x} to bool")
        return None


def lower_and_underscore(s):
    return str(s).strip().lower().replace(' ','_')


def whitespace_to_singlespace(s):
    """Also converts newline characters to single space!"""
    return " ".join(str(s).split())


def remove_reserved_chars(s):
    """Delete Windows file system reserved characters (used if saving a file)"""
    return re.sub('[<>:"/\|?*]', '', str(s))


# =============================================================================
# HTML wrangling
# =============================================================================
def get_text_of_matching_elements(html, xpath_str, re_str=''):
    """Scrape list of cleaned strings from tabular HTML block.

    Intended for use with tabular HTML (e.g., tables, term-description);
    strings have trailing whitespace and characters in re_str removed.
    
    Parameters:
        html: part or all of HTML response
        xpath_str: xpath selector string that may be relative to the given HTML
        remove: string specifying any characters that should be substituted with ''
    """

    data = html.xpath(xpath_str)
    data = data.xpath('string(.)').getall()
    data = [re.sub(re_str, '', d).strip() for d in data]     
    
    return data


def get_text_of_matching_elements_lxml(html, xpath_str, re_str=''):
    """Scrape list of cleaned strings from tabular HTML block.

    Intended for use with tabular HTML (e.g., tables, term-description);
    strings have trailing whitespace and characters in re_str removed.
    
    Parameters:
        html: part or all of HTML response
        xpath_str: xpath selector string that may be relative to the given HTML
        remove: string specifying any characters that should be substituted with ''
    """

    data = html.xpath(xpath_str)
    data = [d.xpath('string(.)') for d in data]
    data = [re.sub(re_str, '', d).strip() for d in data]     
    
    return data


# =============================================================================
# Validate configuration
# =============================================================================
def get_valid_uri(x):
    """Return validated form of the supplied URL or filepath string
    
    Note that on Windows systems, Scrapy needs the 'file' scheme and no netloc"""

    if is_valid_url(x):
        return x
    elif os.path.isfile(x):
        return path_to_uri(filepath)
    else:
        return None


def path_to_uri(filepath):
    """On Windows systems, Scrapy needs the 'file' scheme and no netloc"""
    return 'file:///' + filepath


def is_valid_url(x):
    """Check if a string corresponding to a url is valid
    
    Must contain scheme (e.g., https), netloc (e.g., www.bls.gov), and path"""

    try:
        result = urlparse(x)
        return all([result.scheme, result.netloc, result.path])
    except:
        return False


def get_base_url(x):
    """Get base URL
    
    e.g. if x = https://www.abc.def/ghi/jkl.mno, returns https://www.abc.def"""
    
    result = urlparse(x)
    return urlunparse([result.scheme, result.netloc, '', '', '', ''])


# =============================================================================
# Excel wrangling
# =============================================================================
def load_xls_as_xlsx(filename):
    """Convert xls file to openpyxl xlsx workbook"""

    wb_old = xlrd.open_workbook(filename)
    wb_new =  openpyxl.Workbook()

    for i in range(wb_old.nsheets):
        sheet_old = wb_old.sheet_by_index(i)
        name = wb_old.sheet_names()[i]
        if i == 0:
            sheet_new = wb_new.active
            sheet_new.title = name
        else:
            sheet_new = wb_new.create_sheet(name)

        col_names = []
        for row in range(sheet_old.nrows):
            for col in range(sheet_old.ncols):
                value = sheet_old.cell_value(row, col)
                if row == 0:
                    col_names.append(str(value).lower())
                else:
                    # xlrd imports dates as integers; try to convert to datetime
                    if 'date' in col_names[col]:
                        try:
                            d = xlrd.xldate_as_datetime(int(value), wb_old.datemode)
                            value = d.strftime("%m/%d/%Y")
                        except ValueError:
                            pass
                sheet_new.cell(row=row+1, column=col+1).value = value

    wb_new.save(filename=filename+'x')

    return wb_new
