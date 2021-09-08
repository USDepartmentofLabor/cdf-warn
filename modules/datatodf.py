"""Define 'data-to-DataFrame' (datatodf) functions that extract
pandas DataFrames from tabular HTML, PDF, or Excel files

Each of these functions is content-agnostic (not specific to parsing WARN archives)!"""

from datetime import date
import os
import re
import logging
import camelot
import pandas as pd
import openpyxl
import numpy as np
import bs4
from lxml import etree
from lxml.etree import XMLSyntaxError
from urllib.parse import urljoin

from modules.utils import whitespace_to_singlespace, get_text_of_matching_elements_lxml, load_xls_as_xlsx


# =============================================================================
# Generalized, robust file-to-DataFrame parsing functions
# =============================================================================  
def google_sheets(filepath):
    """Convert Google Sheet (exported as CSV) to DataFrame"""

    return csv(filepath)


def excel(filepath, link_col=None, skip_header=None, drop_footer=None,
          sheets_to_use=None):
    """Convert Excel (.xls or .xslx) sheet to DataFrame
    
    By default, iterates through all worksheets in the workbook,
    using first row of each workbook to generate the column names.

    Parameters:
        link_col: column number or list of column numbers to extract hyperlinks from, if any
        skip_header: number of rows before column names to skip, if any
        drop_footer: number of rows before end of table to skip, if any
        ws_subset: list of worksheet names to parse, if not all
    """

    if skip_header is None:
        skip_header = 0
    if drop_footer is None:
        drop_footer = 0

    # Load workbook, converting old-style Excel to newer version if necessary
    if os.path.splitext(filepath)[-1] == '.xls':
        wb = load_xls_as_xlsx(filepath)
    else:
        wb = openpyxl.load_workbook(filepath)
    
    df_all = pd.DataFrame()

    # Select sheets to parse (all by default, or those specified by sheets_to_use)
    if sheets_to_use:
        worksheets = [wb[i] for i in sheets_to_use]
    else:
        worksheets = wb.worksheets
    for i, ws in enumerate(worksheets):
        df = pd.DataFrame(ws.values)

        # Update column names
        columns = df.iloc[skip_header].to_list()
        columns = [whitespace_to_singlespace(s) for s in columns]
        df.columns = columns

        # Drop unnecessary rows, including column names
        header_rows = list(range(0, skip_header + 1))
        footer_rows = list(range(ws.max_row - drop_footer, ws.max_row))
        df = df.drop(header_rows + footer_rows, axis=0) 

        # Get hidden rows/columns - placeholder

        # Get hyperlinks
        if link_col:
            if isinstance(link_col, int):
                link_col = [link_col]
            for col in link_col:
                # Generate name of new column
                col_name = columns[col]
                if not col_name:
                    col_name = col
                name = col_name + ' Link'

                links = {}
                for row_index in range(ws.min_row + skip_header, ws.max_row - drop_footer): 
                    # Get row_index:link pairs where link exists
                    try:
                        link = ws.cell(row=row_index, col=col).hyperlink.target
                        links[row_index] = link
                    except Exception as e:
                        logging.warning(f"Expecting column {col} to contain a link, but none found")
                        links[row_index] = ''    
                
                df[name] = pd.Series(links).values

        df_all = df_all.append(df)

    return df_all


def pdf(filepath, header_on_all_pages=True, header_row=0, first_page_header=0, pdf_kwargs={},
        column_names=None, col_delimiter=None):
    """Convert PDF consisting of only tables to dataframe
    
    Takes the first table read as representative of the expected number of columns
    and column names for the rest of the tables. Then, subsequent tables are appended 
    to the first if the column numbers match. If a table has fewer colummns than expected,
    it is dropped.

    Parameters:
    - pdf_kwargs:   dictionary of kwargs to pass to Camelot pdf reader
    - header_on_all_pages:  whether each page includes the column names; if not, *Currently in either case. TODO: change?
                            headers are reused from the first table.
    - header_row: number of rows of table to skip
    - column_names: if the column names are for some reason difficult to parse (e.g., poor contrast),
                    provide a list of expected column names, in order
    - col_delimiter:    if the column names are for some reason difficult to parse (e.g., poor contrast),
                        and have been lumped into one line, split them up by the provided delimiter character(s).
    """

    kwargs = {'pages': 'all',
              'flavor': 'lattice',
              'split_text': True} # TODO: See if having this one default helps or hurts
    kwargs.update(pdf_kwargs)
    tables = camelot.read_pdf(filepath, **kwargs)

    frames = []
    df_all = pd.DataFrame()
    for i, table in enumerate(tables):
        df = table.df
        n_cols_current = len(df.columns) 
        
        # Get column names from first page and reuse for remaining pages
        if i == 0:
            if column_names is None:
                column_name_row =  df.iloc[header_row + first_page_header]
                if col_delimiter is None:
                    # If column names pre-specified, use those
                    # Currently used just for CA... may want to split off
                    columns = column_name_row.to_list() 
                else:
                    # If column names get squished into a single column (might not
                    # be first column), split by delimiter
                    # Currently used just for NM... may want to split off
                    index, column_name_str = [(i, c) for (i, c) in enumerate(column_name_row) if c != ''][0]
                    columns = ['']*n_cols_current
                    column_names_partial = column_name_str.split(col_delimiter)
                    columns[index:index+len(column_names_partial)] = column_names_partial
                # Remove whitespace and set column number expectation
                columns = [whitespace_to_singlespace(s) for s in columns]
                n_cols_expected = n_cols_current
            else:
                if n_cols_current == len(column_names):
                    columns = column_names
                    n_cols_expected = len(column_names)
                else:
                    logging.error("The specified column_list does not match the number of columns present; either update column_list or do not provide it")
                    return df_all

        # Some PDFs have summary tables at the end or other table-like elements;
        # column number checks are to discard these. (Keep > for 'stream' flavor)
        if (n_cols_current == n_cols_expected) or (n_cols_current > n_cols_expected and kwargs['flavor'] == 'stream'):
            if (n_cols_current > n_cols_expected) and (kwargs['flavor'] == 'stream'):
                # Drop columns beyond expected columns for 'stream' flavor
                while n_cols_current > n_cols_expected:
                    df.drop(columns=df.columns[-1], axis=1, inplace=True)
                    n_cols_current = len(df.columns)
            df.columns = columns
            
            # Drop rows containing column headings
            # TODO: Check that multi-page pdfs have same header on each page to drop!
            if header_on_all_pages or i == 0:
                if i == 0:
                    rows_to_drop = list(range(0, header_row + first_page_header + 1))
                else:
                    rows_to_drop = list(range(0, header_row + 1))
                df = df.drop(rows_to_drop, axis=0) 

            frames.append(df)
        else:
            logging.warning(f'Dropping a table with {n_cols_current} columns; expected {n_cols_expected} based on first table in file')

    try:
        df_all = pd.concat(frames, join='inner', ignore_index=True) # Drops columns not common to all
    except ValueError as e:
        logging.error(f"{e}: no tables were found. Check if the PDF file was saved correctly?")

    return df_all


def html_pandas(url):   
    """Convert HTML table(s) to DataFrame
    
    If multiple tables are found, they are concatenated into a single DataFrame."""

    try:
        dfs = pd.read_html(url)
    except (XMLSyntaxError, ValueError) as e:
        logging.error(f"{e}. If directly from URL: check if the URL does in fact have HTML tables - if not, update link and/or format. If from temperary file: check if temporary file has any contents - if not, check for error in save_response_to_file.") 
        return pd.DataFrame()

    if len(dfs) > 1:
        try:
            return pd.concat(dfs)
        except Exception as e:
            logging.error(f"{e}: returning empty DataFrame for {url}")
            return pd.DataFrame()
    elif len(dfs) == 1:
        return dfs[0]
    else:
        logging.error(f"No tables found in {url}: returning empty DataFrame")
        return pd.DataFrame()
    

def html(filepath, base_url='', use_pandas=False, tag=None, table_class=None, link_col=None,
         min_num_cols=1, use_unicode=False, drop_first_row=False):
    """Convert HTML table(s) to DataFrame
    
    If multiple tables are found, they are concatenated into a single DataFrame.
    
    While pandas does have a pretty flexible read_html() function, this function
    has some advantages for the purpose of this WARN scraping project:
    - it can parse hyperlinks from a specified column
    - it automatically uses the first rows as the column names if no <th> tags
    
    For other projects, if neither of the above points apply, pd.read_html() is
    probably the way to go, especially if you will directly be using the DataFrame
    for data analysis with pandas.
    """

    # Pandas has a streamlined html table converter that works for many cases;
    # use this by default.
    if use_pandas:
        return html_pandas(filepath)

    if table_class is not None:
        search_str = f'//table[@class="{table_class}"]'
    else:
        if tag is None:
            tag = 'table'
        search_str = f'//{tag}'

    df_all = pd.DataFrame()

    if use_unicode:
        kwargs = {'encoding': 'utf-8'}
    else:
        kwargs = {}

    with open(filepath, 'r', **kwargs) as f:  
        tree = etree.parse(f, etree.HTMLParser())
        tables = tree.xpath(search_str)
        if tables:
            for table in tables:
                # Get column headers, if specified
                columns = get_text_of_matching_elements_lxml(table, './/th')
                
                rows = table.xpath('.//body/tr') # TODO: see if just .//tr is sufficient for all
                if len(rows) == 0:
                    rows = table.xpath('.//tr')
                    if len(rows) == 0:
                        logging.error(f"No table rows found in this table; may need to updated xpath selectors or url")
                        continue

                if drop_first_row:
                    rows.pop(0)

                # If there is no th tag, use first row for column names
                if len(columns) == 0 and len(rows) > 0:
                    columns = get_text_of_matching_elements_lxml(rows.pop(0), './/td')

                for row in rows:
                    # Get value for each column
                    td = get_text_of_matching_elements_lxml(row, './/td')
                    fields = dict(zip(columns, td))

                    # Follow link on each entry to get more detailed information,
                    # updating the original item before yielding it to the pipelines
                    if link_col is not None:
                        #notice_hrefs = row.xpath(f'.//td[{link_col}]/*/@href')
                        notice_hrefs = row.xpath(f'.//*/@href') #TODO: check that this works for all
                        if notice_hrefs:
                            notice_link = urljoin(base_url, notice_hrefs[0])
                            fields.update({'Notice Link': notice_link})
                            if len(notice_hrefs) > 1:
                                fields.update({'Updated Notices': notice_hrefs[1:]})
                        else:
                            logging.warning(f"Expecting a link table entry to contain a link; may need to update column")

                    df_all = df_all.append(fields, ignore_index=True)

        else:
            logging.error(f"No table found; may need to updated xpath selectors or url")

    return df_all


def csv(filepath):
    """Convert HTML table(s) to DataFrame
    
    Trivial, but included for compatibility with config format lookup.
    """

    return pd.read_csv(filepath)

# =============================================================================
# Helper functions
# =============================================================================   
def drop_empty_rows_cols(df):
    """Basic cleaning: remove completely empty rows/columns"""

    # Drop rows and columns if only NaNs
    df.dropna(axis=0, how='all', inplace=True)
    df.dropna(axis=1, how='all', inplace=True)

    # Convert remaining NaNs to empty strings and drop rows/cols consisting of only empty strings 
    df = df.fillna("")
    df = df.loc[:, df.ne("").any()]
    df = df[(df.T != "").any()]

    return df


def reset_spacing_in_column_names(df):
    """Basic cleaning: change newlines and tabsl to single space
    (not yet applied)"""

    # Drop rows and columns if only NaNs
    columns = df.columns
    df.columns = [whitespace_to_singlespace(c) for c in columns]

    return df


def get_google_sheets_export_link(url):
    """for use in spider as:
        response.follow(export_google_sheets_as_excel(url), callback=self.parse_as_df)"""
    
    url_sheet_id = url.split('/edit')[0].strip('/')
    url = url_sheet_id + '/export?format=csv'

    return url   


_ext_dict = {'html': '.html',
             'google_sheets': '.csv',
             'pdf': '.pdf',
             'excel': '.xlsx'}
def get_ext(format):
    ext = _ext_dict.get(format, None)
    if ext is None:
        print("Ya done goofed")
    return ext

    
    